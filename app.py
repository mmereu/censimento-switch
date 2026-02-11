"""Backend Flask per il censimento switch dalla rete NeDi."""

import csv
import io
import ipaddress
import os
import re

import paramiko
import pymysql
import openpyxl
from dotenv import load_dotenv
from flask import Flask, jsonify, render_template, request, send_file

from config import SSH_CONFIG, MYSQL_CONFIG

load_dotenv()

app = Flask(__name__)

DASHBOARD_URL = os.getenv("DASHBOARD_URL", "/")


@app.context_processor
def inject_globals():
    return {"dashboard_url": DASHBOARD_URL}

QUERY_SWITCH = """
SELECT DISTINCT
    d.device AS nome,
    d.vendor AS marca,
    d.type AS modello_raw,
    d.description AS descrizione,
    INET_NTOA(d.devip) AS ip,
    d.serial AS seriale,
    d.location AS posizione,
    d.stack AS n_stack,
    IFNULL(p.porte_up, 0) AS porte_up,
    IFNULL(p.porte_down, 0) AS porte_down
FROM devices d
JOIN networks n ON d.device = n.device
LEFT JOIN (
    SELECT device,
        SUM(CASE WHEN ifstat IN (3, 67, 99) THEN 1 ELSE 0 END) AS porte_up,
        SUM(CASE WHEN ifstat IN (1, 65, 97) THEN 1 ELSE 0 END) AS porte_down
    FROM interfaces
    WHERE ifname LIKE 'Gi%%' OR ifname LIKE 'Eth%%' OR ifname LIKE 'XGi%%'
       OR ifname LIKE 'Te%%' OR ifname LIKE 'GE%%' OR ifname LIKE 'XGE%%'
    GROUP BY device
) p ON d.device = p.device
WHERE INET_NTOA(n.ifip) LIKE %s
ORDER BY d.type, d.device
"""

COLUMNS = ["nome", "marca", "modello", "ip", "seriale", "posizione", "n_stack", "porte_up", "porte_down"]

_MODEL_RE = re.compile(r'\b([A-Z]{1,3}\d{4}[A-Za-z0-9\-]*)')


def _extract_model(raw_type, description):
    """Estrae il modello pulito da type e description NeDi.

    Priorità: ultimo match nella description (modello HW reale alla fine).
    Es: description='...OS (S5700 V600...) ... CloudEngine S5735-L-V2'
    -> S5735-L-V2 (ultimo match, è il modello reale)
    """
    # Se type inizia con un codice modello pulito (es. S5735-L48P4X-A), usalo
    first_word = raw_type.split(" ", 1)[0] if raw_type else ""
    if first_word and re.match(r'^[A-Z]{1,3}\d', first_word):
        # Ma verifica se description ha un modello più specifico
        if description:
            matches = _MODEL_RE.findall(description)
            if matches:
                # Prendi l'ultimo match (il modello HW reale è in fondo)
                last_match = matches[-1]
                if last_match != first_word:
                    return last_match
        return first_word

    # type non ha modello pulito: cerca nella description
    if description:
        matches = _MODEL_RE.findall(description)
        if matches:
            return matches[-1]

    # Ultimo tentativo: cerca in type
    if raw_type:
        matches = _MODEL_RE.findall(raw_type)
        if matches:
            return matches[-1]

    return first_word or raw_type or ""


def _build_like_pattern(network_str, prefix_len):
    """Costruisce il pattern LIKE dalla rete e prefix length.

    Esempi:
        10.31.4.0/24  -> '10.31.4.%'
        10.31.0.0/16  -> '10.31.%'
        10.0.0.0/8    -> '10.%'
    """
    net = ipaddress.ip_network(f"{network_str}/{prefix_len}", strict=False)
    octets = str(net.network_address).split(".")

    if prefix_len >= 24:
        return f"{octets[0]}.{octets[1]}.{octets[2]}.%"
    elif prefix_len >= 16:
        return f"{octets[0]}.{octets[1]}.%"
    elif prefix_len >= 8:
        return f"{octets[0]}.%"
    else:
        return "%"


def _parse_network_params():
    """Estrae e valida i parametri network e prefix dalla request.

    Accetta formati:
        10.31.4.0/24   -> '10.31.4.%'
        10.31.4.0      -> '10.31.4.%'  (default /24)
        10.31.         -> '10.31.%'
        10.31          -> '10.31.%'
    """
    network_raw = request.args.get("network", "").strip().rstrip(".")
    prefix_raw = request.args.get("prefix", "").strip()

    if not network_raw:
        return None, None, "Parametro 'network' mancante"

    # Gestisci formato CIDR (es. 10.31.4.0/24)
    if "/" in network_raw:
        parts = network_raw.split("/", 1)
        network_raw = parts[0]
        prefix_raw = parts[1]

    octets = network_raw.split(".")

    # Input parziale: 1 ottetto (es. "10") -> "10.%"
    # Input parziale: 2 ottetti (es. "10.31") -> "10.31.%"
    # Input parziale: 3 ottetti (es. "10.31.4") -> "10.31.4.%"
    # Input completo: 4 ottetti (es. "10.31.4.0") -> usa prefix
    if len(octets) < 4:
        pattern = ".".join(octets) + ".%"
        prefix_len = len(octets) * 8
        return network_raw, prefix_len, pattern

    # 4 ottetti: validazione completa
    network_str = network_raw
    prefix_str = prefix_raw if prefix_raw else "24"

    try:
        prefix_len = int(prefix_str)
    except ValueError:
        return None, None, f"Prefix non valido: {prefix_str}"

    if prefix_len < 0 or prefix_len > 32:
        return None, None, f"Prefix fuori range (0-32): {prefix_len}"

    try:
        ipaddress.ip_address(network_str)
    except ValueError:
        return None, None, f"Indirizzo IP non valido: {network_str}"

    pattern = _build_like_pattern(network_str, prefix_len)
    return network_str, prefix_len, pattern


def _query_switches(like_pattern):
    """Esegue la query al DB NeDi tramite tunnel SSH con paramiko e restituisce le righe."""
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    pkey = paramiko.RSAKey.from_private_key_file(SSH_CONFIG["ssh_pkey"])
    ssh.connect(
        SSH_CONFIG["ssh_host"],
        port=SSH_CONFIG["ssh_port"],
        username=SSH_CONFIG["ssh_username"],
        pkey=pkey,
    )

    try:
        transport = ssh.get_transport()
        local_port = transport.request_port_forward("", 0)
        # Apri canale diretto verso MySQL sul server remoto
        channel = transport.open_channel(
            "direct-tcpip",
            (MYSQL_CONFIG["host"], MYSQL_CONFIG["port"]),
            ("127.0.0.1", 0),
        )

        conn = pymysql.connect(
            host="127.0.0.1",
            user=MYSQL_CONFIG["user"],
            password=MYSQL_CONFIG["password"],
            database=MYSQL_CONFIG["database"],
            charset=MYSQL_CONFIG["charset"],
            cursorclass=pymysql.cursors.DictCursor,
            defer_connect=True,
        )
        conn.connect(channel)

        try:
            with conn.cursor() as cursor:
                cursor.execute(QUERY_SWITCH, (like_pattern,))
                rows = cursor.fetchall()
        finally:
            conn.close()
    finally:
        ssh.close()

    # Post-process: estrai modello pulito da type + description
    for row in rows:
        row["modello"] = _extract_model(row.pop("modello_raw", ""), row.pop("descrizione", ""))

    return rows


# --- Routes ---


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/search")
def api_search():
    network_str, prefix_len, result = _parse_network_params()
    if network_str is None:
        return jsonify({"error": result}), 400

    try:
        rows = _query_switches(result)
    except Exception as e:
        return jsonify({"error": f"Errore di connessione al database: {e}"}), 500

    return jsonify({
        "network": network_str,
        "prefix": prefix_len,
        "count": len(rows),
        "switches": rows,
    })


@app.route("/api/export/csv")
def api_export_csv():
    network_str, prefix_len, result = _parse_network_params()
    if network_str is None:
        return jsonify({"error": result}), 400

    try:
        rows = _query_switches(result)
    except Exception as e:
        return jsonify({"error": f"Errore di connessione al database: {e}"}), 500

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=COLUMNS)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)

    buf = io.BytesIO(output.getvalue().encode("utf-8"))
    buf.seek(0)

    filename = f"switch_{network_str}_{prefix_len}.csv"
    return send_file(
        buf,
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/export/excel")
def api_export_excel():
    network_str, prefix_len, result = _parse_network_params()
    if network_str is None:
        return jsonify({"error": result}), 400

    try:
        rows = _query_switches(result)
    except Exception as e:
        return jsonify({"error": f"Errore di connessione al database: {e}"}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Switch"

    headers = ["Nome", "Marca", "Modello", "IP", "Seriale", "Posizione", "Stack", "Porte Up", "Porte Down"]
    ws.append(headers)

    for row in rows:
        ws.append([row.get(col, "") for col in COLUMNS])

    # Auto-dimensiona le colonne
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for c in cell:
                if c.value:
                    max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"switch_{network_str}_{prefix_len}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
